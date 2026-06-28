"""
产品参数基准库 CRUD 操作

提供产品线、版本、参数记录的增删改查接口。
"""
import json
import os
import uuid
from datetime import datetime
from typing import Optional

import openpyxl

from config.settings import EXCEL_REQUIRED_FIELDS, EXCEL_FIELD_MAPPING
from database.migrations import get_connection
from database.models import (
    ParameterRecord,
    EvidenceIndex,
    ProductLine,
    ProductVersion,
    ParameterAlias,
)


def _row_to_product_line(row) -> ProductLine:
    return ProductLine(
        id=row["id"],
        name=row["name"],
        description=row["description"] or "",
        created_at=row["created_at"] or "",
    )


def _row_to_product_version(row) -> ProductVersion:
    return ProductVersion(
        id=row["id"],
        product_line_id=row["product_line_id"],
        version_name=row["version_name"],
        release_date=row["release_date"] or "",
        is_active=bool(row["is_active"]),
    )


def _row_to_parameter(row) -> ParameterRecord:
    return ParameterRecord(
        id=row["id"],
        version_id=row["version_id"],
        name=row["name"],
        nominal_value=row["nominal_value"] or "",
        acceptable_range=row["acceptable_range"] or "",
        unit=row["unit"] or "",
        deviation_preset=row["deviation_preset"] or "",
        category=row["category"] or "",
    )


def create_product_line(name: str, description: str = "") -> ProductLine:
    conn = get_connection()
    try:
        pl_id = f"pl_{uuid.uuid4().hex[:12]}"
        conn.execute(
            "INSERT INTO product_lines (id, name, description) VALUES (?, ?, ?)",
            (pl_id, name, description),
        )
        conn.commit()
        return ProductLine(id=pl_id, name=name, description=description)
    finally:
        conn.close()


def get_active_params(product_line_id: str) -> list[ParameterRecord]:
    conn = get_connection()
    try:
        rows = conn.execute(
            """
            SELECT pr.* FROM parameter_records pr
            JOIN product_versions pv ON pr.version_id = pv.id
            WHERE pv.product_line_id = ? AND pv.is_active = 1
            """,
            (product_line_id,),
        ).fetchall()
        return [_row_to_parameter(r) for r in rows]
    finally:
        conn.close()


def get_version_params(version_id: str) -> list[ParameterRecord]:
    conn = get_connection()
    try:
        rows = conn.execute(
            "SELECT * FROM parameter_records WHERE version_id = ?",
            (version_id,),
        ).fetchall()
        return [_row_to_parameter(r) for r in rows]
    finally:
        conn.close()


def get_version_history(product_line_id: str) -> list[ProductVersion]:
    conn = get_connection()
    try:
        rows = conn.execute(
            "SELECT * FROM product_versions WHERE product_line_id = ? ORDER BY release_date DESC",
            (product_line_id,),
        ).fetchall()
        return [_row_to_product_version(r) for r in rows]
    finally:
        conn.close()


def switch_version(product_line_id: str, version_id: str) -> bool:
    conn = get_connection()
    try:
        conn.execute(
            "UPDATE product_versions SET is_active = 0 WHERE product_line_id = ?",
            (product_line_id,),
        )
        conn.execute(
            "UPDATE product_versions SET is_active = 1 WHERE id = ? AND product_line_id = ?",
            (version_id, product_line_id),
        )
        conn.commit()
        return True
    finally:
        conn.close()


def create_product_version(
    product_line_id: str, version_name: str, release_date: str = ""
) -> ProductVersion:
    conn = get_connection()
    try:
        vid = f"v_{uuid.uuid4().hex[:12]}"
        if not release_date:
            release_date = datetime.now().strftime("%Y-%m-%d")
        conn.execute(
            "INSERT INTO product_versions (id, product_line_id, version_name, release_date, is_active) "
            "VALUES (?, ?, ?, ?, 0)",
            (vid, product_line_id, version_name, release_date),
        )
        conn.commit()
        return ProductVersion(
            id=vid,
            product_line_id=product_line_id,
            version_name=version_name,
            release_date=release_date,
            is_active=False,
        )
    finally:
        conn.close()


def import_from_excel(file_path: str, version_id: str = "") -> dict:
    if not os.path.exists(file_path):
        raise FileNotFoundError(f"Excel 文件不存在: {file_path}")

    wb = openpyxl.load_workbook(file_path, data_only=True)
    ws = wb.active

    headers = []
    for cell in ws[1]:
        headers.append(cell.value)

    column_index = {}
    reverse_mapping = {v: k for k, v in EXCEL_FIELD_MAPPING.items()}
    for col_name, field_name in EXCEL_FIELD_MAPPING.items():
        for idx, header in enumerate(headers):
            if header and header.strip() == col_name:
                column_index[field_name] = idx
                break

    missing = [f for f in EXCEL_REQUIRED_FIELDS if EXCEL_FIELD_MAPPING[f] not in column_index]
    if missing:
        raise ValueError(f"Excel 缺少必填字段: {', '.join(missing)}")

    conn = get_connection()
    imported = 0
    errors = []
    try:
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if all(v is None for v in row):
                continue

            name = row[column_index.get("name", -1)]
            nominal_value = row[column_index.get("nominal_value", -1)]
            unit = row[column_index.get("unit", -1)]

            if not name:
                errors.append({"row": row_idx, "error": "参数名称为空"})
                continue

            name = str(name).strip()
            nominal_value = str(nominal_value).strip() if nominal_value else ""
            unit = str(unit).strip() if unit else ""
            acceptable_range = str(row[column_index["acceptable_range"]]).strip() if column_index.get("acceptable_range") is not None and row[column_index["acceptable_range"]] else ""
            deviation_preset = str(row[column_index["deviation_preset"]]).strip() if column_index.get("deviation_preset") is not None and row[column_index["deviation_preset"]] else ""
            category = str(row[column_index["category"]]).strip() if column_index.get("category") is not None and row[column_index["category"]] else ""

            pid = f"p_{uuid.uuid4().hex[:12]}"
            conn.execute(
                "INSERT INTO parameter_records (id, version_id, name, nominal_value, acceptable_range, unit, deviation_preset, category) "
                "VALUES (?, ?, ?, ?, ?, ?, ?, ?)",
                (pid, version_id, name, nominal_value, acceptable_range, unit, deviation_preset, category),
            )

            evidence_title = str(row[column_index["evidence_title"]]).strip() if column_index.get("evidence_title") is not None and len(row) > column_index["evidence_title"] and row[column_index["evidence_title"]] else ""
            if evidence_title:
                evidence_path = str(row[column_index["evidence_path"]]).strip() if column_index.get("evidence_path") is not None and row[column_index["evidence_path"]] else ""
                evidence_page = str(row[column_index["evidence_page_ref"]]).strip() if column_index.get("evidence_page_ref") is not None and row[column_index["evidence_page_ref"]] else ""
                eid = f"e_{uuid.uuid4().hex[:12]}"
                conn.execute(
                    "INSERT INTO evidence_indices (id, parameter_id, doc_title, doc_path, page_ref) VALUES (?, ?, ?, ?, ?)",
                    (eid, pid, evidence_title, evidence_path, evidence_page),
                )

            imported += 1

        conn.commit()
        return {"imported": imported, "errors": errors}
    finally:
        conn.close()


def export_to_excel(version_id: str, output_path: str) -> str:
    params = get_version_params(version_id)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "产品参数"

    col_names = list(EXCEL_FIELD_MAPPING.keys())
    for col_idx, name in enumerate(col_names, 1):
        ws.cell(row=1, column=col_idx, value=name)

    for row_idx, param in enumerate(params, 2):
        field_to_col = {v: k for k, v in enumerate(col_names)}
        ws.cell(row=row_idx, column=field_to_col["参数名称"] + 1, value=param.name)
        ws.cell(row=row_idx, column=field_to_col["标称值"] + 1, value=param.nominal_value)
        ws.cell(row=row_idx, column=field_to_col["可满足范围"] + 1, value=param.acceptable_range)
        ws.cell(row=row_idx, column=field_to_col["单位"] + 1, value=param.unit)
        ws.cell(row=row_idx, column=field_to_col["偏离类型预设"] + 1, value=param.deviation_preset)
        ws.cell(row=row_idx, column=field_to_col["类别"] + 1, value=param.category)

    os.makedirs(os.path.dirname(output_path) or ".", exist_ok=True)
    wb.save(output_path)
    return output_path


def update_parameter(param_id: str, updates: dict) -> Optional[ParameterRecord]:
    conn = get_connection()
    try:
        allowed_fields = [
            "name", "nominal_value", "acceptable_range", "unit",
            "deviation_preset", "category",
        ]
        set_clauses = []
        values = []
        for field in allowed_fields:
            if field in updates:
                set_clauses.append(f"{field} = ?")
                values.append(updates[field])

        if not set_clauses:
            return None

        values.append(param_id)
        conn.execute(
            f"UPDATE parameter_records SET {', '.join(set_clauses)} WHERE id = ?",
            values,
        )
        conn.commit()

        row = conn.execute(
            "SELECT * FROM parameter_records WHERE id = ?", (param_id,)
        ).fetchone()
        return _row_to_parameter(row) if row else None
    finally:
        conn.close()


def add_alias(param_id: str, alias: str) -> bool:
    conn = get_connection()
    try:
        existing = conn.execute(
            "SELECT id FROM parameter_aliases WHERE parameter_id = ? AND alias = ?",
            (param_id, alias),
        ).fetchone()
        if existing:
            return False

        aid = f"a_{uuid.uuid4().hex[:12]}"
        conn.execute(
            "INSERT INTO parameter_aliases (id, parameter_id, alias) VALUES (?, ?, ?)",
            (aid, param_id, alias),
        )
        conn.commit()
        return True
    finally:
        conn.close()


def get_aliases(param_id: str) -> list[str]:
    conn = get_connection()
    try:
        rows = conn.execute(
            "SELECT alias FROM parameter_aliases WHERE parameter_id = ?",
            (param_id,),
        ).fetchall()
        return [r["alias"] for r in rows]
    finally:
        conn.close()


def delete_alias(param_id: str, alias: str) -> bool:
    conn = get_connection()
    try:
        cursor = conn.execute(
            "DELETE FROM parameter_aliases WHERE parameter_id = ? AND alias = ?",
            (param_id, alias),
        )
        conn.commit()
        return cursor.rowcount > 0
    finally:
        conn.close()
